import openpyxl as xl
import cv2
import pytesseract
import pdfplumber
import docx

# I installed pytesseract on my computer from the following: https://github.com/UB-Mannheim/tesseract/wiki
# Mine is 64-bit
# Point the below line to where your tesseract.exe file is located. I installed mine in the filepath below
pytesseract.pytesseract.tesseract_cmd = r"C:\\Program Files\\Tesseract-OCR\\tesseract.exe"


# Because openxl does not support csv files, I first copied the cell-content in the file '1. Data for the task 1.csv'
# and pasted it into a an excel spreadsheet and saved it as an '.xlsx' file. I then ran the code that I made below to
# convert male and female values into 1 and 0 respectively. I want to be able to use the gender column in Jupyter when
# I create my 'X' dataset. Before loading it into Jupyter, I then converted the '.xlsx' file back to '.csv' in order
# for me to use read_csv() to create my dataframe


def load_spreadsheet(spreadsheet, sheet, starting_row):
    wb = xl.load_workbook(spreadsheet)
    sheet = wb[sheet]
    starting_row = starting_row

    for row in range(starting_row, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, column)
            if column == 3:
                if cell.value == 'F':
                    new_cell_value = 0
                elif cell.value == 'M':
                    new_cell_value = 1
                cell.value = new_cell_value

    wb.save('../Data_Task1-converted.xlsx')


# load_spreadsheet('../Data_Task1.xlsx', 'Sheet1', 2)


class ReadContents:
    def __init__(self, filename):
        self.filename = filename

    def read_image(self):
        image = cv2.imread(self.filename)
        text = pytesseract.image_to_string(image)
        return text

    def read_pdf(self):
        with pdfplumber.open(self.filename) as pdf:
            first_page = pdf.pages[0]
            text = first_page.extract_text()
            return text

    def read_docx(self):
        doc = docx.Document(self.filename)
        full_text = []
        for para in doc.paragraphs:
            full_text.append(para.text)
        text = '\n'.join(full_text)
        return text


# a = ReadContents('../media/documents/sample.jpg')
# text = a.read_image()
# print(text)